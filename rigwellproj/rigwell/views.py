# import datetime
import difflib
import os
import string
import urllib
from itertools import islice

import io
import requests
import xlrd
import re

from django.core import mail
from django.core.mail import send_mail, BadHeaderError, EmailMessage
from django.contrib import messages
# from _mysql_exceptions import DataError, IntegrityError
from django.template import RequestContext

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.core.mail import EmailMultiAlternatives

from django.core.files.storage import FileSystemStorage
import json
from django.contrib import auth
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect, HttpResponseNotAllowed
from django.utils.datastructures import MultiValueDictKeyError
from django.views.decorators.cache import cache_control
from numpy import long
from openpyxl.styles import PatternFill

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.fields import empty
from rest_framework.permissions import AllowAny
from time import gmtime, strftime
import time
from xlrd import XLRDError

from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.contrib.sessions.backends.db import SessionStore
from django.contrib.sessions.models import Session
from django.contrib.auth.models import User, AnonymousUser
from django.conf import settings
from django import forms
import sys
from django.core.cache import cache
import random

from rigwell.models import RWMember, Rig, Well, Activity, TimeSDL, TDExtended, TDTIH, TDDrlg, TDPOH, TimeSDLField, TDExtendedField, TDTIHField, TDDrlgField, TDPOHField

class UploadFileForm(forms.Form):
    file = forms.FileField()

def index(request):
    return HttpResponse('<h1>Hello! I am RigWell Website.</h1>')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rig_admin(request):
    return render(request, 'rigwell/rig_admin.html')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rig_admin_login(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        name = request.POST.get('name', '')

        members = RWMember.objects.filter(email='hafiz.kurnia92@gmail.com', role='admin')
        # if members.count() == 0:
        #     member = RWMember()
        #     member.name = name
        #     member.email = email
        #     member.password = 'admin!@#$%'
        #     member.role = 'admin'
        #     member.save()
        #     request.session['adminID'] = member.pk
        #     return redirect('/rahome')
        if members.count() > 0:
            member = members[0]
            request.session['adminID'] = member.pk
            return redirect('/rahome')
        else:
            return render(request, 'rigwell/rig_result.html',
                          {'response': 'You don\'t have any permission to access this site. Try again with another credential.'})


def rig_admin_home(request):
    import datetime
    try:
        if request.session['adminID'] == 0 or request.session['adminID'] == '':
            return render(request, 'rigwell/rig_admin.html')
    except KeyError:
        print('no key')
        return render(request, 'rigwell/rig_admin.html')

    adminID = request.session['adminID']
    admin = RWMember.objects.get(id=adminID)

    rigs = Rig.objects.filter(member_id=adminID)
    rigList = []
    for rig in rigs:
        wells = Well.objects.filter(rig_id=rig.pk)
        wellList = []
        for well in wells:
            well.well_start_time = datetime.datetime.fromtimestamp(float(int(well.well_start_time)/1000)).strftime("%b %d, %Y %H:%M")
            well.well_end_time = datetime.datetime.fromtimestamp(float(int(well.well_end_time)/1000)).strftime("%b %d, %Y %H:%M")

            activities = Activity.objects.filter(well_id=well.pk)
            for activity in activities:
                activity.time_start = datetime.datetime.fromtimestamp(float(int(activity.time_start)/1000)).strftime("%b %d, %Y %H:%M")
                activity.time_end = datetime.datetime.fromtimestamp(float(int(activity.time_end)/1000)).strftime("%b %d, %Y %H:%M")
            data = {
                'well':well,
                'activities':activities
            }
            wellList.append(data)
        data = {
            'rig':rig,
            'wells':wellList,
        }
        rigList.append(data)

    selrigid = 0
    selwellid = 0
    selactid = 0

    try:
        selrigid = request.GET['selrigid']
    except KeyError:
        print('No key')

    try:
        selwellid = request.GET['selwellid']
    except KeyError:
        print('No key')

    try:
        selactid = request.GET['selactid']
    except KeyError:
        print('No key')

    return render(request, 'rigwell/rig_admin_home.html', {'me':admin, 'datas':rigList, 'selrigid':int(selrigid), 'selwellid':int(selwellid), 'selactid':int(selactid)})


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rig_admin_new_rig(request):
    if request.method == 'POST':

        name = request.POST.get('rig_name', '')
        type = request.POST.get('rig_type', '')
        contractor = request.POST.get('rig_contractor', '')

        try:
            if request.session['adminID'] == 0 or request.session['adminID'] == '':
                return render(request, 'rigwell/rig_admin.html')
        except KeyError:
            print('no key')
            return render(request, 'rigwell/rig_admin.html')

        adminID = request.session['adminID']
        admin = RWMember.objects.get(id=adminID)

        fs = FileSystemStorage()

        rig = Rig()
        rig.member_id = admin.pk
        rig.name = name
        rig.type = type
        rig.contractor = contractor
        rig.created_time = str(int(round(time.time() * 1000)))

        try:
            image = request.FILES['rig_image']
            filename = fs.save(image.name, image)
            uploaded_url = fs.url(filename)
            rig.picture_url = settings.URL + uploaded_url
        except MultiValueDictKeyError:
            print('no video updated')

        rig.save()

        return redirect('/rahome?selrigid=' + str(rig.pk))



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rig_admin_new_well(request):
    import datetime
    if request.method == 'POST':

        rig_id = request.POST.get('rig_id', '1')
        name = request.POST.get('well_name', '')
        field = request.POST.get('well_field', '')
        country = request.POST.get('well_country', '')
        start_time = request.POST.get('well_start_time', '')
        end_time = request.POST.get('well_end_time', '')

        start_time = str(int(datetime.datetime.strptime(start_time, "%Y/%m/%d %H:%M").timestamp()*1000))
        end_time = str(int(datetime.datetime.strptime(end_time, "%Y/%m/%d %H:%M").timestamp()*1000))

        try:
            if request.session['adminID'] == 0 or request.session['adminID'] == '':
                return render(request, 'rigwell/rig_admin.html')
        except KeyError:
            print('no key')
            return render(request, 'rigwell/rig_admin.html')

        adminID = request.session['adminID']
        admin = RWMember.objects.get(id=adminID)

        well = Well()
        well.rig_id = rig_id
        well.name = name
        well.field = field
        well.country = country
        well.well_start_time = start_time
        well.well_end_time = end_time

        well.save()

        return redirect('/rahome?selwellid=' + str(well.pk))




@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rig_admin_new_activity(request):
    import datetime
    if request.method == 'POST':

        well_id = request.POST.get('well_id', '1')
        name = request.POST.get('act_name', '')
        number = request.POST.get('act_number', '0')
        hole = request.POST.get('act_hole', '')
        time_start = request.POST.get('time_start', '')
        time_end = request.POST.get('time_end', '')
        depth_start = request.POST.get('act_depth_start', '0')
        depth_end = request.POST.get('act_depth_end', '0')

        time_start = str(int(datetime.datetime.strptime(time_start, "%Y/%m/%d %H:%M").timestamp()*1000))
        time_end = str(int(datetime.datetime.strptime(time_end, "%Y/%m/%d %H:%M").timestamp()*1000))

        try:
            if request.session['adminID'] == 0 or request.session['adminID'] == '':
                return render(request, 'rigwell/rig_admin.html')
        except KeyError:
            print('no key')
            return render(request, 'rigwell/rig_admin.html')

        adminID = request.session['adminID']
        admin = RWMember.objects.get(id=adminID)

        act = Activity()
        act.well_id = well_id
        act.name = name
        act.number = number
        act.hole = hole
        act.time_start = time_start
        act.time_end = time_end
        act.depth_start = depth_start
        act.depth_end = depth_end
        act.total_rows = '1'
        act.saved_rows = '0'

        act.save()

        return redirect('/rahome?selactid=' + str(act.pk))



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rig_admin_act_item_setup(request):
    if request.method == 'POST':

        act_id = request.POST.get('act_id', '1')
        item = request.POST.get('act_item', '')

        try:
            if request.session['adminID'] == 0 or request.session['adminID'] == '':
                return render(request, 'rigwell/rig_admin.html')
        except KeyError:
            print('no key')
            return render(request, 'rigwell/rig_admin.html')

        adminID = request.session['adminID']
        admin = RWMember.objects.get(id=adminID)

        request.session['act_id'] = act_id
        request.session['act_item'] = item

        acts = Activity.objects.filter(id=act_id)
        if acts.count() == 0:
            return redirect('/rahome')
        act = acts[0]
        wells = Well.objects.filter(id=act.well_id)
        if wells.count() == 0:
            return redirect('/rahome')
        well = wells[0]

        rigs = Rig.objects.filter(id=well.rig_id)
        if rigs.count() == 0:
            return redirect('/rahome')
        rig = rigs[0]

        if item == 'Time SDL':
            fields = TimeSDLField.objects.all()
            return render(request, 'rigwell/rig_admin_activity_set_data.html', {'me':admin, 'act':act, 'well':well, 'rig':rig, 'item':item, 'fields':fields})
        elif item == 'TD Extended':
            fields = TDExtendedField.objects.all()
            return render(request, 'rigwell/rig_admin_activity_set_data.html', {'me':admin, 'act':act, 'well':well, 'rig':rig, 'item':item, 'fields':fields})
        elif item == 'TD TIH Stat':
            fields = TDTIHField.objects.all()
            return render(request, 'rigwell/rig_admin_activity_set_data.html', {'me':admin, 'act':act, 'well':well, 'rig':rig, 'item':item, 'fields':fields})
        elif item == 'TD Drlg Stat':
            fields = TDDrlgField.objects.all()
            return render(request, 'rigwell/rig_admin_activity_set_data.html', {'me':admin, 'act':act, 'well':well, 'rig':rig, 'item':item, 'fields':fields})
        elif item == 'TD POH Stat':
            fields = TDPOHField.objects.all()
            return render(request, 'rigwell/rig_admin_activity_set_data.html', {'me':admin, 'act':act, 'well':well, 'rig':rig, 'item':item, 'fields':fields})
        elif item == 'Time/Depth':
            return render(request, 'rigwell/rig_admin_activity_set_data.html', {'me':admin, 'act':act, 'well':well, 'rig':rig, 'item':item, 'fields':fields})
        elif item == 'All Items':
            return redirect('/raimportexcel')
        else:
            return redirect('/rahome')

    else:

        act_id = request.session['act_id']
        item = request.session['act_item']

        try:
            if request.session['adminID'] == 0 or request.session['adminID'] == '':
                return render(request, 'rigwell/rig_admin.html')
        except KeyError:
            print('no key')
            return render(request, 'rigwell/rig_admin.html')

        adminID = request.session['adminID']
        admin = RWMember.objects.get(id=adminID)

        acts = Activity.objects.filter(id=act_id)
        if acts.count() == 0:
            return redirect('/rahome')
        act = acts[0]
        wells = Well.objects.filter(id=act.well_id)
        if wells.count() == 0:
            return redirect('/rahome')
        well = wells[0]

        rigs = Rig.objects.filter(id=well.rig_id)
        if rigs.count() == 0:
            return redirect('/rahome')
        rig = rigs[0]

        if item == 'Time SDL':
            fields = TimeSDLField.objects.all()
            return render(request, 'rigwell/rig_admin_activity_set_data.html', {'me':admin, 'act':act, 'well':well, 'rig':rig, 'item':item, 'fields':fields})
        elif item == 'TD Extended':
            fields = TDExtendedField.objects.all()
            return render(request, 'rigwell/rig_admin_activity_set_data.html', {'me':admin, 'act':act, 'well':well, 'rig':rig, 'item':item, 'fields':fields})
        elif item == 'TD TIH Stat':
            fields = TDTIHField.objects.all()
            return render(request, 'rigwell/rig_admin_activity_set_data.html', {'me':admin, 'act':act, 'well':well, 'rig':rig, 'item':item, 'fields':fields})
        elif item == 'TD Drlg Stat':
            fields = TDDrlgField.objects.all()
            return render(request, 'rigwell/rig_admin_activity_set_data.html', {'me':admin, 'act':act, 'well':well, 'rig':rig, 'item':item, 'fields':fields})
        elif item == 'TD POH Stat':
            fields = TDPOHField.objects.all()
            return render(request, 'rigwell/rig_admin_activity_set_data.html', {'me':admin, 'act':act, 'well':well, 'rig':rig, 'item':item, 'fields':fields})
        elif item == 'Time/Depth':
            return render(request, 'rigwell/rig_admin_activity_set_data.html', {'me':admin, 'act':act, 'well':well, 'rig':rig, 'item':item, 'fields':fields})
        elif item == 'All Items':
            return redirect('/raimportexcel')
        else:
            return redirect('/rahome')



def rig_admin_logout(request):
    request.session['adminID'] = 0
    return redirect('/admin')


def rig_admin_import_excel(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)
    else:
        form = UploadFileForm()

    act_id = None
    item = ''
    try:
        act_id = request.session['act_id']
        item = request.session['act_item']
    except KeyError:
        print('No key')

    context = {
        'form': form,
        'title': 'Load Data',
        'header': 'Upload Data From File'
    }

    if act_id is not None:
        acts = Activity.objects.filter(id=act_id)
        if acts.count() == 0:
            return redirect('/rahome')
        act = acts[0]
        wells = Well.objects.filter(id=act.well_id)
        if wells.count() == 0:
            return redirect('/rahome')
        well = wells[0]

        rigs = Rig.objects.filter(id=well.rig_id)
        if rigs.count() == 0:
            return redirect('/rahome')
        rig = rigs[0]

        context = {
            'form': form,
            'title': 'Load Data',
            'header': 'Upload Data From File',
            'act':act, 'well':well, 'rig':rig, 'item':item
        }

    return render(
        request, 'rigwell/rig_admin_excel_upload.html', context)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rig_admin_import_excel_data(request):

    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)

        try:
            if request.session['adminID'] == 0:
                return render(request, 'rigwell/rig_admin.html')
        except KeyError:
            print('no session')
            return render(request, 'rigwell/rig_admin.html')

        adminID = request.session['adminID']

        act_id = request.session['act_id']
        item = request.session['act_item']

        acts = Activity.objects.filter(id=act_id)
        if acts.count() == 0:
            return redirect('/rahome')
        act = acts[0]
        wells = Well.objects.filter(id=act.well_id)
        if wells.count() == 0:
            return redirect('/rahome')
        well = wells[0]

        rigs = Rig.objects.filter(id=well.rig_id)
        if rigs.count() == 0:
            return redirect('/rahome')
        rig = rigs[0]

        timesdl_fields = TimeSDL._meta.fields
        tdextended_fields = TDExtended._meta.fields
        tdtih_fields = TDTIH._meta.fields
        tddrlg_fields = TDDrlg._meta.fields
        tdpoh_fields = TDPOH._meta.fields

        fieldList = [timesdl_fields, tdextended_fields, tdtih_fields, tddrlg_fields, tdpoh_fields]
        itemNameList = ['Time SDL', 'TD Extended', 'TD TIH Stat', 'TD Drlg Stat', 'TD POH Stat']
        threshold = 500000

        act.total_rows = '1'
        act.saved_rows = '0'
        act.save()

        if form.is_valid():
            input_excel = request.FILES['file']
            try:
                book = xlrd.open_workbook(file_contents=input_excel.read())
                list = []
                if item == 'All Items':
                    ss = len(book.sheet_names())
                    total_rows = 0
                    for s in range(1, ss):
                        sheet = book.sheet_by_index(s)
                        total_rows = total_rows + sheet.nrows

                    act.total_rows = str(total_rows)
                    act.save()

                    loaded_rows = 0

                    for s in range(1, ss):   # from the second sheet
                        sheet = book.sheet_by_index(s)
                        rs = threshold
                        if sheet.nrows < rs:
                            rs = sheet.nrows

                        si = itemNameList.index(book.sheet_names()[s])

                        for r in range(1, rs):  # from the second row (values only)
                            obj = None
                            if si == 0:
                                obj = TimeSDL()
                            elif si == 1:
                                obj = TDExtended()
                            elif si == 2:
                                obj = TDTIH()
                            elif si == 3:
                                obj = TDDrlg()
                            elif si == 4:
                                obj = TDPOH()

                            obj.rig_id = rig.pk
                            obj.well_id = well.pk
                            obj.act_id = act.pk

                            i = 0   # from the first column
                            for field in fieldList[si]:
                                field_name = field.name
                                if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                                    val = sheet.cell(r, i).value
                                    if val is not None:
                                        if isinstance(val, int):
                                            val = str(val).replace('.0', '')
                                        elif isinstance(val, float):
                                            val = str(round(val, 5))
                                        setattr(obj, field_name, val)
                                        i = i + 1
                            obj.save()

                            loaded_rows = loaded_rows + 1
                            act.saved_rows = str(loaded_rows)
                            act.save()

                        list.append(itemNameList[si] + ' is success!<br><br>')
                else:
                    if not item in book.sheet_names():
                        return HttpResponse('no_sheet')

                    sheet_index = book.sheet_names().index(item)
                    sheet = book.sheet_by_index(sheet_index)

                    act.total_rows = str(sheet.nrows)
                    act.save()

                    loaded_rows = 0

                    rs = threshold
                    if sheet.nrows < rs:
                        rs = sheet.nrows

                    si = itemNameList.index(item)

                    for r in range(1, rs):  # from the second row (values only)
                        obj = None
                        if si == 0:
                            obj = TimeSDL()
                        elif si == 1:
                            obj = TDExtended()
                        elif si == 2:
                            obj = TDTIH()
                        elif si == 3:
                            obj = TDDrlg()
                        elif si == 4:
                            obj = TDPOH()

                        obj.rig_id = rig.pk
                        obj.well_id = well.pk
                        obj.act_id = act.pk

                        i = 0   # from the first column
                        for field in fieldList[si]:
                            field_name = field.name
                            if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                                val = sheet.cell(r, i).value
                                if val is not None:
                                    if isinstance(val, int):
                                        val = str(val).replace('.0', '')
                                    elif isinstance(val, float):
                                        val = str(round(val, 5))
                                    setattr(obj, field_name, val)
                                    i = i + 1
                        obj.save()

                        loaded_rows = loaded_rows + 1
                        act.saved_rows = str(loaded_rows)
                        act.save()

                    list.append(itemNameList[si] + ' is success!<br><br>')

                return HttpResponse(list)

            except XLRDError:
                return render(request, 'rigwell/rig_admin_excel_upload.html', {'note': 'invalid_file'})
            except IOError:
                return render(request, 'rigwell/rig_admin_excel_upload.html', {'note': 'invalid_file'})
            except IndexError:
                return render(request, 'rigwell/rig_admin_excel_upload.html', {'note': 'invalid_file'})
            # except DataError:
            #     return HttpResponse('Invalid file!')
        else:
            return render(request, 'rigwell/rig_admin_excel_upload.html', {'note': 'invalid_file'})


def rig_admin_get_loaded_excel_total_rows(request):

    act_id = request.session['act_id']
    acts = Activity.objects.filter(id=act_id)
    if acts.count() == 0:
        return HttpResponse(0)

    act = acts[0]
    total_rows = int(act.total_rows)
    saved_rows = int(act.saved_rows)

    return HttpResponse(float(saved_rows*100/total_rows))


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rig_admin_act_manually_setup_data(request):
    if request.method == 'POST':

        act_id = request.POST.get('act_id', '1')
        item = request.POST.get('item', '')

        acts = Activity.objects.filter(id=act_id)
        if acts.count() == 0:
            return redirect('/rahome')
        act = acts[0]
        wells = Well.objects.filter(id=act.well_id)
        if wells.count() == 0:
            return redirect('/rahome')
        well = wells[0]

        rigs = Rig.objects.filter(id=well.rig_id)
        if rigs.count() == 0:
            return redirect('/rahome')
        rig = rigs[0]

        try:
            if request.session['adminID'] == 0 or request.session['adminID'] == '':
                return render(request, 'rigwell/rig_admin.html')
        except KeyError:
            print('no key')
            return render(request, 'rigwell/rig_admin.html')

        adminID = request.session['adminID']
        admin = RWMember.objects.get(id=adminID)

        vals = request.POST.getlist('values[]')

        timesdl_fields = TimeSDL._meta.fields
        tdextended_fields = TDExtended._meta.fields
        tdtih_fields = TDTIH._meta.fields
        tddrlg_fields = TDDrlg._meta.fields
        tdpoh_fields = TDPOH._meta.fields

        fieldList = [timesdl_fields, tdextended_fields, tdtih_fields, tddrlg_fields, tdpoh_fields]
        itemNameList = ['Time SDL', 'TD Extended', 'TD TIH Stat', 'TD Drlg Stat', 'TD POH Stat']

        index = itemNameList.index(item)

        obj = None
        if index == 0:
            obj = TimeSDL()
        elif index == 1:
            obj = TDExtended()
        elif index == 2:
            obj = TDTIH()
        elif index == 3:
            obj = TDDrlg()
        elif index == 4:
            obj = TDPOH()

        obj.rig_id = rig.pk
        obj.well_id = well.pk
        obj.act_id = act.pk

        vallength = len(vals)
        for i in range(0, vallength):
            val = vals[i]
            field = fieldList[index][i+4]
            field_name = field.name
            if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                if val is not None:
                    if isinstance(val, int):
                        val = str(val).replace('.0', '')
                    elif isinstance(val, float):
                        val = str(round(val, 5))
                    setattr(obj, field_name, val)

        obj.save()

        # list = []
        # for val in vals:
        #     list.append(val)
        # return HttpResponse(list)


        timesdl_fields = TimeSDL._meta.fields
        tdextended_fields = TDExtended._meta.fields
        tdtih_fields = TDTIH._meta.fields
        tddrlg_fields = TDDrlg._meta.fields
        tdpoh_fields = TDPOH._meta.fields

        list = []
        fields = []
        r = None

        if index == 0:
            objs = TimeSDL.objects.filter(act_id=act_id).order_by('-id')
            objs, r = get_activity_data(objs)
            for obj in objs:
                valList = []
                for field in timesdl_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TimeSDLField.objects.all()
        elif index == 1:
            objs = TDExtended.objects.filter(act_id=act_id).order_by('-id')
            objs, r = get_activity_data(objs)
            for obj in objs:
                valList = []
                for field in tdextended_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TDExtendedField.objects.all()
        elif index == 2:
            objs = TDTIH.objects.filter(act_id=act_id).order_by('-id')
            objs, r = get_activity_data(objs)
            for obj in objs:
                valList = []
                for field in tdtih_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TDTIHField.objects.all()
        elif index == 3:
            objs = TDDrlg.objects.filter(act_id=act_id).order_by('-id')
            objs, r = get_activity_data(objs)
            for obj in objs:
                valList = []
                for field in tddrlg_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TDDrlgField.objects.all()
        elif index == 4:
            objs = TDPOH.objects.filter(act_id=act_id).order_by('-id')
            objs, r = get_activity_data(objs)
            for obj in objs:
                valList = []
                for field in tdpoh_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TDPOHField.objects.all()

        return render(request, 'rigwell/rig_admin_activity_data.html', {'me':admin, 'act':act, 'well':well, 'rig':rig, 'item':item, 'datalist':list, 'range':r, 'current':1, 'fields':fields})





def rig_admin_view_activity_data(request):

    act_id = None
    item = ''
    try:
        act_id = request.session['act_id']
        item = request.session['act_item']
    except KeyError:
        print('No key')

    try:
        if request.session['adminID'] == 0 or request.session['adminID'] == '':
            return render(request, 'rigwell/rig_admin.html')
    except KeyError:
        print('no key')
        return render(request, 'rigwell/rig_admin.html')

    adminID = request.session['adminID']
    admin = RWMember.objects.get(id=adminID)

    if act_id is not None:
        acts = Activity.objects.filter(id=act_id)
        if acts.count() == 0:
            return redirect('/rahome')
        act = acts[0]
        wells = Well.objects.filter(id=act.well_id)
        if wells.count() == 0:
            return redirect('/rahome')
        well = wells[0]

        rigs = Rig.objects.filter(id=well.rig_id)
        if rigs.count() == 0:
            return redirect('/rahome')
        rig = rigs[0]

        itemNameList = ['Time SDL', 'TD Extended', 'TD TIH Stat', 'TD Drlg Stat', 'TD POH Stat']
        if item == 'All Items': index = 0
        else: index = itemNameList.index(item)

        timesdl_fields = TimeSDL._meta.fields
        tdextended_fields = TDExtended._meta.fields
        tdtih_fields = TDTIH._meta.fields
        tddrlg_fields = TDDrlg._meta.fields
        tdpoh_fields = TDPOH._meta.fields

        list = []
        fields = []
        range = None

        if index == 0:
            objs = TimeSDL.objects.filter(act_id=act_id).order_by('-id')
            objs, range = get_activity_data(objs)
            for obj in objs:
                valList = []
                for field in timesdl_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TimeSDLField.objects.all()
        elif index == 1:
            objs = TDExtended.objects.filter(act_id=act_id).order_by('-id')
            objs, range = get_activity_data(objs)
            for obj in objs:
                valList = []
                for field in tdextended_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TDExtendedField.objects.all()
        elif index == 2:
            objs = TDTIH.objects.filter(act_id=act_id).order_by('-id')
            objs, range = get_activity_data(objs)
            for obj in objs:
                valList = []
                for field in tdtih_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TDTIHField.objects.all()
        elif index == 3:
            objs = TDDrlg.objects.filter(act_id=act_id).order_by('-id')
            objs, range = get_activity_data(objs)
            for obj in objs:
                valList = []
                for field in tddrlg_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TDDrlgField.objects.all()
        elif index == 4:
            objs = TDPOH.objects.filter(act_id=act_id).order_by('-id')
            objs, range = get_activity_data(objs)
            for obj in objs:
                valList = []
                for field in tdpoh_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TDPOHField.objects.all()

        return render(request, 'rigwell/rig_admin_activity_data.html', {'me':admin, 'act':act, 'well':well, 'rig':rig, 'item':item, 'datalist':list, 'range':range, 'current':1, 'fields':fields})

    else:
        return redirect('/rahome')



def get_activity_data(items):
    i = 0
    itemList = []
    for item in items:
        i = i + 1
        if i <= 25:
            itemList.append(item)
    r = int(len(items) / 25)
    m = len(items) % 25
    if m > 0:
        r = r + 2
    else:
        r = r + 1
    return itemList, range(r)


def ra_to_page(request):
    idx = request.GET['index']

    try:
        if request.session['adminID'] == 0 or request.session['adminID'] == '':
            return render(request, 'rigwell/rig_admin.html')
    except KeyError:
        print('no key')
        return render(request, 'rigwell/rig_admin.html')

    adminID = request.session['adminID']
    admin = RWMember.objects.get(id=adminID)

    if int(idx) == 1:
        return redirect('/raviewactivitydata')

    act_id = None
    item = ''
    try:
        act_id = request.session['act_id']
        item = request.session['act_item']
    except KeyError:
        print('No key')

    try:
        if request.session['adminID'] == 0 or request.session['adminID'] == '':
            return render(request, 'rigwell/rig_admin.html')
    except KeyError:
        print('no key')
        return render(request, 'rigwell/rig_admin.html')

    adminID = request.session['adminID']
    admin = RWMember.objects.get(id=adminID)

    if act_id is not None:
        acts = Activity.objects.filter(id=act_id)
        if acts.count() == 0:
            return redirect('/rahome')
        act = acts[0]
        wells = Well.objects.filter(id=act.well_id)
        if wells.count() == 0:
            return redirect('/rahome')
        well = wells[0]

        rigs = Rig.objects.filter(id=well.rig_id)
        if rigs.count() == 0:
            return redirect('/rahome')
        rig = rigs[0]

        itemNameList = ['Time SDL', 'TD Extended', 'TD TIH Stat', 'TD Drlg Stat', 'TD POH Stat']
        if item == 'All Items': index = 0
        else: index = itemNameList.index(item)

        timesdl_fields = TimeSDL._meta.fields
        tdextended_fields = TDExtended._meta.fields
        tdtih_fields = TDTIH._meta.fields
        tddrlg_fields = TDDrlg._meta.fields
        tdpoh_fields = TDPOH._meta.fields

        list = []
        fields = []
        range = None

        if index == 0:
            objs = TimeSDL.objects.filter(act_id=act_id).order_by('-id')
            objs, range = get_to_page_data(objs, idx)
            for obj in objs:
                valList = []
                for field in timesdl_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TimeSDLField.objects.all()
        elif index == 1:
            objs = TDExtended.objects.filter(act_id=act_id).order_by('-id')
            objs, range = get_to_page_data(objs, idx)
            for obj in objs:
                valList = []
                for field in tdextended_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TDExtendedField.objects.all()
        elif index == 2:
            objs = TDTIH.objects.filter(act_id=act_id).order_by('-id')
            objs, range = get_to_page_data(objs, idx)
            for obj in objs:
                valList = []
                for field in tdtih_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TDTIHField.objects.all()
        elif index == 3:
            objs = TDDrlg.objects.filter(act_id=act_id).order_by('-id')
            objs, range = get_to_page_data(objs, idx)
            for obj in objs:
                valList = []
                for field in tddrlg_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TDDrlgField.objects.all()
        elif index == 4:
            objs = TDPOH.objects.filter(act_id=act_id).order_by('-id')
            objs, range = get_to_page_data(objs, idx)
            for obj in objs:
                valList = []
                for field in tdpoh_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TDPOHField.objects.all()

        return render(request, 'rigwell/rig_admin_activity_data.html', {'me':admin, 'act':act, 'well':well, 'rig':rig, 'item':item, 'datalist':list, 'range':range, 'current':idx, 'fields':fields})

    else:
        return redirect('/rahome')


def get_to_page_data(items, index):

    i = 0
    itemList = []
    for item in items:
        i = i + 1
        if i > 25 * (int(index) - 1) and i <= 25 * int(index):
            itemList.append(item)
    r = int(len(items) / 25)
    r = r + 2
    return itemList, range(r)


def ra_to_previous(request):
    idx = request.GET['index']

    try:
        if request.session['adminID'] == 0 or request.session['adminID'] == '':
            return render(request, 'rigwell/rig_admin.html')
    except KeyError:
        print('no key')
        return render(request, 'rigwell/rig_admin.html')

    adminID = request.session['adminID']
    admin = RWMember.objects.get(id=adminID)

    if int(idx) == 1:
        return redirect('/raviewactivitydata')

    index = int(idx) - 1
    return redirect('/ra_to_page?index=' + str(index))


def ra_to_next(request):
    idx = request.GET['index']

    try:
        if request.session['adminID'] == 0 or request.session['adminID'] == '':
            return render(request, 'rigwell/rig_admin.html')
    except KeyError:
        print('no key')
        return render(request, 'rigwell/rig_admin.html')

    adminID = request.session['adminID']
    admin = RWMember.objects.get(id=adminID)

    if int(idx) == 1:
        return redirect('/raviewactivitydata')

    act_id = None
    item = ''
    try:
        act_id = request.session['act_id']
        item = request.session['act_item']
    except KeyError:
        print('No key')

    try:
        if request.session['adminID'] == 0 or request.session['adminID'] == '':
            return render(request, 'rigwell/rig_admin.html')
    except KeyError:
        print('no key')
        return render(request, 'rigwell/rig_admin.html')

    adminID = request.session['adminID']
    admin = RWMember.objects.get(id=adminID)

    if act_id is not None:
        acts = Activity.objects.filter(id=act_id)
        if acts.count() == 0:
            return redirect('/rahome')
        act = acts[0]
        wells = Well.objects.filter(id=act.well_id)
        if wells.count() == 0:
            return redirect('/rahome')
        well = wells[0]

        rigs = Rig.objects.filter(id=well.rig_id)
        if rigs.count() == 0:
            return redirect('/rahome')
        rig = rigs[0]

        itemNameList = ['Time SDL', 'TD Extended', 'TD TIH Stat', 'TD Drlg Stat', 'TD POH Stat']
        if item == 'All Items': index = 0
        else: index = itemNameList.index(item)

        if index == 0:
            objs = TimeSDL.objects.filter(act_id=act_id)
            idx = get_next_page_data(objs, idx)
            return redirect('/ra_to_page?index=' + str(idx))
        elif index == 1:
            objs = TDExtended.objects.filter(act_id=act_id)
            idx = get_next_page_data(objs, idx)
            return redirect('/ra_to_page?index=' + str(idx))
        elif index == 2:
            objs = TDTIH.objects.filter(act_id=act_id)
            idx = get_next_page_data(objs, idx)
            return redirect('/ra_to_page?index=' + str(idx))
        elif index == 3:
            objs = TDDrlg.objects.filter(act_id=act_id)
            idx = get_next_page_data(objs, idx)
            return redirect('/ra_to_page?index=' + str(idx))
        elif index == 4:
            objs = TDPOH.objects.filter(act_id=act_id)
            idx = get_next_page_data(objs, idx)
            return redirect('/ra_to_page?index=' + str(idx))

        return redirect('/raviewactivitydata')

    else:
        return redirect('/rahome')


def get_next_page_data(items, index):
    count = len(items)
    r = int(count / 25)
    m = count % 25
    if m > 0:
        r = r + 2
    else:
        r = r + 1
    if int(index) < r - 1:
        index = int(index) + 1

    return index




def rig_admin_activity_data_view(request):

    request.session['url'] = settings.URL + request.get_full_path()

    try:
        if request.session['adminID'] == 0 or request.session['adminID'] == '':
            return render(request, 'rigwell/rig_admin.html')
    except KeyError:
        print('no key')
        return render(request, 'rigwell/rig_admin.html')

    adminID = request.session['adminID']
    admin = RWMember.objects.get(id=adminID)

    index = request.GET['idx']
    index = int(index)
    act_id = request.GET['act_id']
    well_id = request.GET['well_id']
    rig_id = request.GET['rig_id']

    itemNameList = ['Time SDL', 'TD Extended', 'TD TIH Stat', 'TD Drlg Stat', 'TD POH Stat']
    item = itemNameList[index]

    request.session['act_id'] = act_id
    request.session['act_item'] = item

    if act_id is not None:
        acts = Activity.objects.filter(id=act_id)
        if acts.count() == 0:
            return redirect('/rahome')
        act = acts[0]
        wells = Well.objects.filter(id=well_id)
        if wells.count() == 0:
            return redirect('/rahome')
        well = wells[0]

        rigs = Rig.objects.filter(id=rig_id)
        if rigs.count() == 0:
            return redirect('/rahome')
        rig = rigs[0]

        timesdl_fields = TimeSDL._meta.fields
        tdextended_fields = TDExtended._meta.fields
        tdtih_fields = TDTIH._meta.fields
        tddrlg_fields = TDDrlg._meta.fields
        tdpoh_fields = TDPOH._meta.fields

        list = []
        fields = []
        range = None

        if index == 0:
            objs = TimeSDL.objects.filter(act_id=act_id).order_by('-id')
            objs, range = get_activity_data(objs)
            for obj in objs:
                valList = []
                for field in timesdl_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TimeSDLField.objects.all()
        elif index == 1:
            objs = TDExtended.objects.filter(act_id=act_id).order_by('-id')
            objs, range = get_activity_data(objs)
            for obj in objs:
                valList = []
                for field in tdextended_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TDExtendedField.objects.all()
        elif index == 2:
            objs = TDTIH.objects.filter(act_id=act_id).order_by('-id')
            objs, range = get_activity_data(objs)
            for obj in objs:
                valList = []
                for field in tdtih_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TDTIHField.objects.all()
        elif index == 3:
            objs = TDDrlg.objects.filter(act_id=act_id).order_by('-id')
            objs, range = get_activity_data(objs)
            for obj in objs:
                valList = []
                for field in tddrlg_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TDDrlgField.objects.all()
        elif index == 4:
            objs = TDPOH.objects.filter(act_id=act_id).order_by('-id')
            objs, range = get_activity_data(objs)
            for obj in objs:
                valList = []
                for field in tdpoh_fields:
                    field_name = field.name
                    if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                        valList.append(getattr(obj, field_name, ''))
                data={
                    'values':valList
                }
                list.append(data)
            fields = TDPOHField.objects.all()

        return render(request, 'rigwell/rig_admin_activity_data.html', {'me':admin, 'act':act, 'well':well, 'rig':rig, 'item':item, 'datalist':list, 'range':range, 'current':1, 'fields':fields})

    else:
        return redirect('/rahome')


def rig_admin_delete_rig(request):
    rig_id = request.GET['rig_id']

    rigs = Rig.objects.filter(id=rig_id)
    if rigs.count() == 0:
        return render(request, 'rigwell/rig_result.html',
                          {'response': 'The rig doesn\'t exist.'})
    rig = rigs[0]
    wells = Well.objects.filter(rig_id=rig.pk)
    for well in wells:
        activities = Activity.objects.filter(well_id=well.pk)
        for activity in activities:
            datas = TimeSDL.objects.filter(rig_id=rig.pk, well_id=well.pk, act_id=activity.pk)
            for data in datas:
                data.delete()

            datas = TDExtended.objects.filter(rig_id=rig.pk, well_id=well.pk, act_id=activity.pk)
            for data in datas:
                data.delete()

            datas = TDTIH.objects.filter(rig_id=rig.pk, well_id=well.pk, act_id=activity.pk)
            for data in datas:
                data.delete()

            datas = TDDrlg.objects.filter(rig_id=rig.pk, well_id=well.pk, act_id=activity.pk)
            for data in datas:
                data.delete()

            datas = TDPOH.objects.filter(rig_id=rig.pk, well_id=well.pk, act_id=activity.pk)
            for data in datas:
                data.delete()

            activity.delete()

        well.delete()
    rig.delete()

    return redirect('/rahome')


def rig_admin_delete_well(request):
    well_id = request.GET['well_id']

    wells = Well.objects.filter(id=well_id)
    if wells.count() == 0:
        return render(request, 'rigwell/rig_result.html',
                          {'response': 'The well doesn\'t exist.'})
    well = wells[0]

    activities = Activity.objects.filter(well_id=well.pk)
    for activity in activities:
        datas = TimeSDL.objects.filter(rig_id=int(well.rig_id), well_id=well.pk, act_id=activity.pk)
        for data in datas:
            data.delete()

        datas = TDExtended.objects.filter(rig_id=int(well.rig_id), well_id=well.pk, act_id=activity.pk)
        for data in datas:
            data.delete()

        datas = TDTIH.objects.filter(rig_id=int(well.rig_id), well_id=well.pk, act_id=activity.pk)
        for data in datas:
            data.delete()

        datas = TDDrlg.objects.filter(rig_id=int(well.rig_id), well_id=well.pk, act_id=activity.pk)
        for data in datas:
            data.delete()

        datas = TDPOH.objects.filter(rig_id=int(well.rig_id), well_id=well.pk, act_id=activity.pk)
        for data in datas:
            data.delete()

        activity.delete()

    well.delete()

    return redirect('/rahome')


def rig_admin_delete_activity(request):
    act_id = request.GET['act_id']

    activities = Activity.objects.filter(id=act_id)
    if activities.count() == 0:
        return render(request, 'rigwell/rig_result.html',
                          {'response': 'The activity doesn\'t exist.'})
    activity = activities[0]

    datas = TimeSDL.objects.filter(act_id=activity.pk)
    for data in datas:
        data.delete()

    datas = TDExtended.objects.filter(act_id=activity.pk)
    for data in datas:
        data.delete()

    datas = TDTIH.objects.filter(act_id=activity.pk)
    for data in datas:
        data.delete()

    datas = TDDrlg.objects.filter(act_id=activity.pk)
    for data in datas:
        data.delete()

    datas = TDPOH.objects.filter(act_id=activity.pk)
    for data in datas:
        data.delete()

    activity.delete()

    return redirect('/rahome')


def rig_admin_clear_activity_item_data(request):
    act_id = request.GET['act_id']
    item = request.GET['item']

    activities = Activity.objects.filter(id=act_id)
    if activities.count() == 0:
        return render(request, 'rigwell/rig_result.html',
                          {'response': 'The activity doesn\'t exist.'})
    activity = activities[0]

    itemNameList = ['Time SDL', 'TD Extended', 'TD TIH Stat', 'TD Drlg Stat', 'TD POH Stat']
    index = itemNameList.index(item)

    if index == 0:
        datas = TimeSDL.objects.filter(act_id=activity.pk)
        for data in datas:
            data.delete()
    elif index == 1:
        datas = TDExtended.objects.filter(act_id=activity.pk)
        for data in datas:
            data.delete()
    elif index == 2:
        datas = TDTIH.objects.filter(act_id=activity.pk)
        for data in datas:
            data.delete()
    elif index == 3:
        datas = TDDrlg.objects.filter(act_id=activity.pk)
        for data in datas:
            data.delete()
    elif index == 4:
        datas = TDPOH.objects.filter(act_id=activity.pk)
        for data in datas:
            data.delete()

    try:
        return redirect(request.session['url'])
    except KeyError:
        return redirect('/rahome')
    return redirect('/rahome')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rig_admin_edit_rig(request):
    if request.method == 'POST':

        rig_id = request.POST.get('rig_id', '1')
        name = request.POST.get('rig_name', '')
        type = request.POST.get('rig_type', '')
        contractor = request.POST.get('rig_contractor', '')

        try:
            if request.session['adminID'] == 0 or request.session['adminID'] == '':
                return render(request, 'rigwell/rig_admin.html')
        except KeyError:
            print('no key')
            return render(request, 'rigwell/rig_admin.html')

        adminID = request.session['adminID']
        admin = RWMember.objects.get(id=adminID)

        fs = FileSystemStorage()

        rigs = Rig.objects.filter(id=rig_id)
        if rigs.count() == 0:
            return render(request, 'rigwell/rig_result.html', {'response':'The rig doesn\'t exist'})
        rig = rigs[0]
        rig.name = name
        rig.type = type
        rig.contractor = contractor

        try:
            image = request.FILES['rig_image']
            filename = fs.save(image.name, image)
            uploaded_url = fs.url(filename)
            rig.picture_url = settings.URL + uploaded_url
        except MultiValueDictKeyError:
            print('no video updated')

        rig.save()

        return redirect('/rahome?selrigid=' + str(rig.pk))



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rig_admin_edit_well(request):
    import datetime
    if request.method == 'POST':

        well_id = request.POST.get('well_id', '1')
        name = request.POST.get('well_name', '')
        field = request.POST.get('well_field', '')
        country = request.POST.get('well_country', '')
        start_time = request.POST.get('well_start_time', '')
        end_time = request.POST.get('well_end_time', '')

        if start_time != '': start_time = str(int(datetime.datetime.strptime(start_time, "%Y/%m/%d %H:%M").timestamp()*1000))
        if end_time != '': end_time = str(int(datetime.datetime.strptime(end_time, "%Y/%m/%d %H:%M").timestamp()*1000))

        try:
            if request.session['adminID'] == 0 or request.session['adminID'] == '':
                return render(request, 'rigwell/rig_admin.html')
        except KeyError:
            print('no key')
            return render(request, 'rigwell/rig_admin.html')

        adminID = request.session['adminID']
        admin = RWMember.objects.get(id=adminID)

        wells = Well.objects.filter(id=well_id)
        if wells.count() == 0:
            return render(request, 'rigwell/rig_result.html', {'response':'The well doesn\'t exist'})
        well = wells[0]
        well.name = name
        well.field = field
        well.country = country
        if start_time != '': well.well_start_time = start_time
        if end_time != '': well.well_end_time = end_time

        well.save()

        return redirect('/rahome?selwellid=' + str(well.pk))


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rig_admin_edit_activity(request):
    import datetime
    if request.method == 'POST':

        act_id = request.POST.get('act_id', '1')
        name = request.POST.get('act_name', '')
        number = request.POST.get('act_number', '0')
        hole = request.POST.get('act_hole', '')
        time_start = request.POST.get('time_start', '')
        time_end = request.POST.get('time_end', '')
        depth_start = request.POST.get('act_depth_start', '0')
        depth_end = request.POST.get('act_depth_end', '0')

        if time_start != '': time_start = str(int(datetime.datetime.strptime(time_start, "%Y/%m/%d %H:%M").timestamp()*1000))
        if time_end != '': time_end = str(int(datetime.datetime.strptime(time_end, "%Y/%m/%d %H:%M").timestamp()*1000))

        try:
            if request.session['adminID'] == 0 or request.session['adminID'] == '':
                return render(request, 'rigwell/rig_admin.html')
        except KeyError:
            print('no key')
            return render(request, 'rigwell/rig_admin.html')

        adminID = request.session['adminID']
        admin = RWMember.objects.get(id=adminID)

        acts = Activity.objects.filter(id=act_id)
        if acts.count() == 0:
            return render(request, 'rigwell/rig_result.html', {'response':'The activity doesn\'t exist'})
        act = acts[0]
        act.name = name
        act.number = number
        act.hole = hole
        if time_start != '': act.time_start = time_start
        if time_end != '': act.time_end = time_end
        act.depth_start = depth_start
        act.depth_end = depth_end

        act.save()

        return redirect('/rahome?selactid=' + str(act.pk))


#################################################################### RIGWELL Activity Item Data Fields Set Up ###########################################################################

def rig_admin_set_fields(request):
    timesdl_fields = TimeSDL._meta.fields
    tdextended_fields = TDExtended._meta.fields
    tdtih_fields = TDTIH._meta.fields
    tddrlg_fields = TDDrlg._meta.fields
    tdpoh_fields = TDPOH._meta.fields

    fieldList = [timesdl_fields, tdextended_fields, tdtih_fields, tddrlg_fields, tdpoh_fields]
    for f in range(0, 5):
        for field in fieldList[f]:
            field_name = field.name
            if field.name != 'id' and field.name != 'rig_id' and field.name != 'well_id' and field.name != 'act_id' and field.name != 'status':
                if f == 0:
                    obj = TimeSDL.objects.first()
                    timesdlfield = TimeSDLField()
                    timesdlfield.field_name = getattr(obj, field_name, '')
                    timesdlfield.save()
                elif f == 1:
                    obj = TDExtended.objects.first()
                    tdextendedfield = TDExtendedField()
                    tdextendedfield.field_name = getattr(obj, field_name, '')
                    tdextendedfield.save()
                elif f == 2:
                    obj = TDTIH.objects.first()
                    tdtihfield = TDTIHField()
                    tdtihfield.field_name = getattr(obj, field_name, '')
                    tdtihfield.save()
                elif f == 3:
                    obj = TDDrlg.objects.first()
                    tddrlgfield = TDDrlgField()
                    tddrlgfield.field_name = getattr(obj, field_name, '')
                    tddrlgfield.save()
                elif f == 4:
                    obj = TDPOH.objects.first()
                    tdpohfield = TDPOHField()
                    tdpohfield.field_name = getattr(obj, field_name, '')
                    tdpohfield.save()

    return HttpResponse('DONE!!!')




################################################ RIGWELL User ###############################################################################################################

def rig_user_loginpage(request):
    try:
        if request.session['memberID'] != 0 and request.session['memberID'] != '':
            return redirect('/ruhome')
    except KeyError:
        print('no session')
    return render(request, 'rigwell/rig_user_login.html')


def rig_user_signuppage(request):
    return render(request, 'rigwell/rig_user_signup.html')


def rig_user_forgotpasswordpage(request):
    return render(request, 'rigwell/rig_user_forgotpassword.html')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rig_user_login(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')

        members = RWMember.objects.filter(email=email, password=password)
        if members.count() > 0:
            member = members[0]
            if member.role == 'admin':
                return render(request, 'rigwell/rig_result.html',
                          {'response': 'This account isn\'t allowed to login as a member.'})

            request.session['memberID'] = member.pk
            return redirect('/ruhome')
        else:
            members = RWMember.objects.filter(email=email)
            if members.count() > 0:
                member = members[0]
                if member.role == 'admin':
                    return render(request, 'rigwell/rig_result.html',
                              {'response': 'This account isn\'t allowed to login as a member.'})
                else:
                    return render(request, 'rigwell/rig_result.html',
                              {'response': 'Your password is incorrect. Try again with your correct password.'})
            else:
                return render(request, 'rigwell/rig_result.html',
                          {'response': 'This account doesn\'t exist. Please sign up.'})

    else:
        return redirect('/ruhome')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rig_user_signup(request):
    if request.method == 'POST':
        name = request.POST.get('name', '')
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')

        members = RWMember.objects.filter(email=email, role='')
        count = members.count()
        if count ==0:
            member = RWMember()
            member.name = name
            member.email = email
            member.password = password
            member.role = ''
            member.save()

            request.session['memberID'] = member.pk

            return redirect('/ruhome')

        else:
            return redirect('/rulogout')


def rig_user_home(request):
    import datetime
    try:
        if request.session['memberID'] == 0 or request.session['memberID'] == '':
            return render(request, 'rigwell/rig_user_login.html')
    except KeyError:
        print('no session')
        return render(request, 'rigwell/rig_user_login.html')

    memberID = request.session['memberID']
    me = RWMember.objects.get(id=memberID)

    rigs = Rig.objects.all().order_by('-id')
    rigList = []
    for rig in rigs:
        wells = Well.objects.filter(rig_id=rig.pk).order_by('-id')
        wellList = []
        for well in wells:
            well.well_start_time = datetime.datetime.fromtimestamp(float(int(well.well_start_time)/1000)).strftime("%d-%b-%y,%H:%M")
            well.well_end_time = datetime.datetime.fromtimestamp(float(int(well.well_end_time)/1000)).strftime("%d-%b-%y,%H:%M")

            activities = Activity.objects.filter(well_id=well.pk).order_by('-id')
            for activity in activities:
                activity.time_start = datetime.datetime.fromtimestamp(float(int(activity.time_start)/1000)).strftime("%d-%b-%y,%H:%M")
                activity.time_end = datetime.datetime.fromtimestamp(float(int(activity.time_end)/1000)).strftime("%d-%b-%y,%H:%M")
            wdata = {
                'well':well,
                'activities':activities
            }
            wellList.append(wdata)
        data = {
            'rig':rig,
            'wells':wellList,
        }
        rigList.append(data)

    if memberID == 2: return render(request, 'rigwell/rig_zzz_home.html', {'me':me, 'datas':rigList})

    return render(request, 'rigwell/rig_user_home.html', {'me':me, 'datas':rigList})


def rig_user_logout(request):
    request.session['memberID'] = 0
    return redirect('/user')


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rig_user_send_mail_forgotpassword(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')

        members = RWMember.objects.filter(email=email)
        if members.count() == 0:
            return render(request, 'rigwell/rig_result.html',
                          {'response': 'This email doesn\'t exist. Please try another one.'})

        message = 'You are allowed to reset your password from your request.<br>For it, please click this link to reset your password.<br><br><a href=\'' + settings.URL + '/ruresetpassword?email=' + email
        message = message + '\' target=\'_blank\'>' + 'Link to reset password' + '</a>'

        html =  """\
                    <html>
                        <head></head>
                        <body>
                            <a href="#"><img src="https://hafizkurnia.pythonanywhere.com/static/images/rig_logo.jpg" style="width:120px; height:120px; border-radius:10px; margin-left:25px;"/></a>
                            <h2 style="color:#02839a;">THK User's Security Update Information</h2>
                            <div style="font-size:14px; word-break: break-all; word-wrap: break-word;">
                                {mes}
                            </div>
                        </body>
                    </html>
                """
        html = html.format(mes=message)

        fromEmail = 'hafiz.kurnia92@gmail.com'
        toEmailList = []
        toEmailList.append(email)
        msg = EmailMultiAlternatives('We allowed you to reset your password', '', fromEmail, toEmailList)
        msg.attach_alternative(html, "text/html")
        msg.send(fail_silently=False)

        return render(request, 'rigwell/rig_result.html',
                          {'response': 'We sent a message to your email. Please check and reset your password.'})


def rig_user_resetpassword(request):
    email = request.GET['email']
    return render(request, 'rigwell/rig_user_resetpwd.html', {'email':email})


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rig_user_rstpwd(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')

        members = RWMember.objects.filter(email=email)
        if members.count() == 0:
            return render(request, 'rigwell/rig_result.html',
                          {'response': 'This email doesn\'t exist.'})

        member = members[0]
        member.password = password
        member.save()

        return render(request, 'rigwell/rig_user_login.html', {'notify':'password changed'})



def rig_user_passwordchange(request):
    try:
        if request.session['memberID'] == 0 or request.session['memberID'] == '':
            return render(request, 'rigwell/rig_user_login.html')
    except KeyError:
        print('no session')
        return render(request, 'rigwell/rig_user_login.html')

    memberID = request.session['memberID']
    me = RWMember.objects.get(id=memberID)

    return  render(request, 'rigwell/rig_user_password_reset.html', {'me':me})



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rig_user_changepassword(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        oldpassword = request.POST.get('oldpassword', '')
        newpassword = request.POST.get('newpassword', '')

        try:
            if request.session['memberID'] == 0 or request.session['memberID'] == '':
                return render(request, 'rigwell/rig_user_login.html')
        except KeyError:
            print('no session')
            return render(request, 'rigwell/rig_user_login.html')

        memberID = request.session['memberID']
        me = RWMember.objects.get(id=memberID)

        if email == me.email and oldpassword == me.password:
            me.password = newpassword

            me.save()

        elif email == me.email and oldpassword != me.password:
            return render(request, 'rigwell/rig_result.html',
                          {'response': 'Your old password is incorrect. Please enter your correct password.'})

        else:
            return render(request, 'rigwell/rig_result.html',
                          {'response': 'Your email or password is incorrect. Please enter your correct information.'})

        return  render(request, 'rigwell/rig_user_password_reset.html', {'me':me, 'note':'password_updated'})



def rig_user_well_detail(request):
    well_id = request.GET['well_id']
    try:
        if request.session['memberID'] == 0 or request.session['memberID'] == '':
            return render(request, 'rigwell/rig_user_login.html')
    except KeyError:
        print('no session')
        return render(request, 'rigwell/rig_user_login.html')

    memberID = request.session['memberID']
    me = RWMember.objects.get(id=memberID)

    wells = Well.objects.filter(id=well_id)
    if wells.count() == 0:
        return redirect('/ruhome')
    well = wells[0]

    rigs = Rig.objects.filter(id=well.rig_id)
    if rigs.count() == 0:
        return redirect('/ruhome')
    rig = rigs[0]

    activities = Activity.objects.filter(well_id=well.pk)

    depthDataList = []
    holeDepthDataList = []

    timeDateList = []
    depthList = []
    depthList2 = []
    holeDepthList = []
    chartWidth = 0
    chartHeight = 0

    phlist1 = []
    phlist2 = []
    phlist3 = []
    phlist4 = []
    phlist5 = []
    phlist6 = []
    phlist7 = []

    actHoles = [
        '26 Inch Section',
        '17-1/2 Inch Section',
        '12-1/4 Inch Section',
        '8-1/2 Inch Section',
        '6-1/2 Inch Phase',
        'Completion Phase',
        'Plug and Abandonment Phase'
    ]

    phLists = [
        phlist1,
        phlist2,
        phlist3,
        phlist4,
        phlist5,
        phlist6,
        phlist7
    ]

    for activity in activities:
        objs = TimeSDL.objects.filter(act_id=activity.pk)
        last_timestamp = 0
        for obj in objs:
            import datetime
            timestamp = int(datetime.datetime.strptime(obj.a, "%d-%b-%y %H:%M:%S.%f").timestamp()*1000)
            if timestamp >= last_timestamp + 3600 * 12 * 1000:   # pick data every 12 hours
                last_timestamp = timestamp
                timeDateList.append(obj.a)
                from datetime import datetime
                dtObj = datetime.fromtimestamp(timestamp/1000)
                depth = {
                    'year': dtObj.year,
                    'month': dtObj.month - 1,
                    'day': dtObj.day,
                    'hour': dtObj.hour,
                    'minute': dtObj.minute,
                    'second': dtObj.second,
                    'depth': float(obj.b),
                }
                depthDataList.append(depth)

                hole_depth = {
                    'year': dtObj.year,
                    'month': dtObj.month - 1,
                    'day': dtObj.day,
                    'hour': dtObj.hour,
                    'minute': dtObj.minute,
                    'second': dtObj.second,
                    'depth': float(obj.d),
                }
                holeDepthDataList.append(hole_depth)

                depthList.append(obj.b)
                depthList2.append(str(float(obj.b)*(-1)))
                holeDepthList.append(obj.d)
                chartHeight = chartHeight + 20

        data = {
            'phase_section': activity.hole,
            'activity_name': activity.name,
            'start_hole_depth': objs[0].d,
            'end_hole_depth': objs[objs.count()-1].d,
            'distance_drilled': str(abs(round(float(objs[objs.count()-1].d) - float(objs[0].d), 5))),
            'start_time': objs[0].a,
            'end_time': objs[objs.count()-1].a,
            'duration': getDurationStr(int(datetime.datetime.strptime(objs[objs.count()-1].a, "%d-%b-%y %H:%M:%S.%f").timestamp()*1000)\
                - int(datetime.datetime.strptime(objs[0].a, "%d-%b-%y %H:%M:%S.%f").timestamp()*1000))
        }

        phLists[actHoles.index(activity.hole)].append(data)

    import datetime
    well.well_start_time = datetime.datetime.fromtimestamp(float(int(well.well_start_time)/1000)).strftime("%d-%b-%y,%H:%M")
    well.well_end_time = datetime.datetime.fromtimestamp(float(int(well.well_end_time)/1000)).strftime("%d-%b-%y,%H:%M")

    context = {
        'me':me,
        'rig':rig,
        'well':well,
        'depthDataList':depthDataList,
        'holeDepthDataList':holeDepthDataList,
        'timedates':timeDateList,
        'depths':depthList,
        'depths2':depthList2,
        'holedepths':holeDepthList,
        'chartHeight':chartHeight,
        'datas_1':{'data':phLists[0], 'phase_section':actHoles[0]},
        'datas_2':{'data':phLists[1], 'phase_section':actHoles[1]},
        'datas_3':{'data':phLists[2], 'phase_section':actHoles[2]},
        'datas_4':{'data':phLists[3], 'phase_section':actHoles[3]},
        'datas_5':{'data':phLists[4], 'phase_section':actHoles[4]},
        'datas_6':{'data':phLists[5], 'phase_section':actHoles[5]},
        'datas_7':{'data':phLists[6], 'phase_section':actHoles[6]},
        'opt':'well_detail'
    }

    if memberID == 2:
        return  render(request, 'rigwell/rig_zzz_well_detail.html', context)

    return  render(request, 'rigwell/rig_user_well_detail.html', context)



def getDurationStr(timestamp):
    yearmilis = 365*86400*1000
    monthmilis = 30*86400*1000
    daymilis = 86400*1000
    hourmilis = 3600*1000
    minutemilis = 60*1000
    secondmilis = 1000

    years = int(timestamp / yearmilis)
    months = int((timestamp % yearmilis) / monthmilis)
    days = int(((timestamp % yearmilis) % monthmilis) / daymilis)
    hours = int((((timestamp % yearmilis) % monthmilis) % daymilis) / hourmilis)
    minutes = int(((((timestamp % yearmilis) % monthmilis) % daymilis) % hourmilis) / minutemilis)
    seconds = int((((((timestamp % yearmilis) % monthmilis) % daymilis) % hourmilis) % minutemilis) / secondmilis)

    duration = ''
    if years > 0: duration = str(years) + ' yrs '
    if months > 0: duration = duration + str(months) + ' mths '
    if days > 0: duration = duration + str(days) + ' days '
    if hours > 0: duration = duration + str(hours) + ' hrs '
    if minutes > 0: duration = duration + str(minutes) + ' mins '

    return duration


def rig_user_eow(request):

    well_id = request.GET['well_id']
    try:
        if request.session['memberID'] == 0 or request.session['memberID'] == '':
            return render(request, 'rigwell/rig_user_login.html')
    except KeyError:
        print('no session')
        return render(request, 'rigwell/rig_user_login.html')

    memberID = request.session['memberID']
    me = RWMember.objects.get(id=memberID)

    wells = Well.objects.filter(id=well_id)
    if wells.count() == 0:
        return redirect('/ruhome')
    well = wells[0]

    rigs = Rig.objects.filter(id=well.rig_id)
    if rigs.count() == 0:
        return redirect('/ruhome')
    rig = rigs[0]

    activities = Activity.objects.filter(well_id=well.pk)

    depthDataList = []
    holeDepthDataList = []

    timeDateList = []
    depthList = []
    depthList2 = []
    holeDepthList = []
    chartWidth = 0
    chartHeight = 0

    phlist1 = []
    phlist2 = []
    phlist3 = []
    phlist4 = []
    phlist5 = []
    phlist6 = []
    phlist7 = []

    actHoles = [
        '26 Inch Section',
        '17-1/2 Inch Section',
        '12-1/4 Inch Section',
        '8-1/2 Inch Section',
        '6-1/2 Inch Phase',
        'Completion Phase',
        'Plug and Abandonment Phase'
    ]

    phLists = [
        phlist1,
        phlist2,
        phlist3,
        phlist4,
        phlist5,
        phlist6,
        phlist7
    ]

    inteprActs = ['TIH BHA Conn', 'TIH Connection', 'TIH BHA', 'TIH String', 'TIH Circ', 'TIH Ream Down', 'TIH Ream Up', 'TIH Wash Down', 'TIH Wash Up', 'Drlg Rotary', 'Drlg Sliding', \
        'Drlg Circ', 'Drlg Circ Post Conn', 'Drlg Connection', 'POH BHA Conn', 'POH Connection', 'POH BHA', 'POH String', 'POH Circ', 'POH Ream Down', 'POH Ream Up', 'POH Wash Down', 'POH Wash Up', 'Drlg Ream Down', \
        'Drlg Ream Up', 'Drlg Wash Down', 'Drlg Wash Up', 'Unknown']

    inteprActValFields = ['v','r','u','q','z','ad','ac','ab','aa','at','au','ax','aw','av','aj','af','ai','ae','ao','a_s','ar','aq','ap','bx','bw','bv','bu','bq']

    inteprActVals = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    inteprActPercents = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]

    total = 0

    for activity in activities:
        objs = TimeSDL.objects.filter(act_id=activity.pk)
        last_timestamp = 0
        for obj in objs:
            import datetime
            timestamp = int(datetime.datetime.strptime(obj.a, "%d-%b-%y %H:%M:%S.%f").timestamp()*1000)
            if timestamp >= last_timestamp + 3600 * 12 * 1000:   # pick data every an hour
                last_timestamp = timestamp
                timeDateList.append(obj.a)
                from datetime import datetime
                dtObj = datetime.fromtimestamp(timestamp/1000)
                depth = {
                    'year': dtObj.year,
                    'month': dtObj.month - 1,
                    'day': dtObj.day,
                    'hour': dtObj.hour,
                    'minute': dtObj.minute,
                    'second': dtObj.second,
                    'depth': float(obj.b),
                }
                depthDataList.append(depth)

                hole_depth = {
                    'year': dtObj.year,
                    'month': dtObj.month - 1,
                    'day': dtObj.day,
                    'hour': dtObj.hour,
                    'minute': dtObj.minute,
                    'second': dtObj.second,
                    'depth': float(obj.d),
                }
                holeDepthDataList.append(hole_depth)

                depthList.append(obj.b)
                depthList2.append(str(float(obj.b)*(-1)))
                holeDepthList.append(obj.d)
                chartHeight = chartHeight + 20

        data = {
            'phase_section': activity.hole,
            'activity_name': activity.name,
            'start_hole_depth': objs[0].d,
            'end_hole_depth': objs[objs.count()-1].d,
            'distance_drilled': str(abs(round(float(objs[objs.count()-1].d) - float(objs[0].d), 5))),
            'start_time': objs[0].a,
            'end_time': objs[objs.count()-1].a,
            'duration': getDurationStr(int(datetime.datetime.strptime(objs[objs.count()-1].a, "%d-%b-%y %H:%M:%S.%f").timestamp()*1000)\
                - int(datetime.datetime.strptime(objs[0].a, "%d-%b-%y %H:%M:%S.%f").timestamp()*1000))
        }

        phLists[actHoles.index(activity.hole)].append(data)

        objs = TDExtended.objects.filter(act_id=activity.pk)
        for obj in objs:
            if obj.e in inteprActs:
                index = inteprActs.index(obj.e)
                field_name = inteprActValFields[index]
                val = getattr(obj, field_name)
                inteprActVals[index] = inteprActVals[index] + float(val)
                total = total + float(val)

    timeBreakdownList = []
    timeBreakdownSubList1 = []
    timeBreakdownSubList2 = []

    if total > 0:
        for i in range(0, len(inteprActVals)):
            inteprActPercents[i] = round(inteprActVals[i] * 100 / total, 2)
            data = {
                'title':inteprActs[i],
                'value':inteprActPercents[i]
            }
            timeBreakdownList.append(data)
            if i < len(inteprActVals)/2:
                timeBreakdownSubList1.append(data)
            else:
                timeBreakdownSubList2.append(data)

    import datetime
    well.well_start_time = datetime.datetime.fromtimestamp(float(int(well.well_start_time)/1000)).strftime("%d-%b-%y,%H:%M")
    well.well_end_time = datetime.datetime.fromtimestamp(float(int(well.well_end_time)/1000)).strftime("%d-%b-%y,%H:%M")

    context = {
        'me':me,
        'rig':rig,
        'well':well,
        'depthDataList':depthDataList,
        'holeDepthDataList':holeDepthDataList,
        'timedates':timeDateList,
        'depths':depthList,
        'depths2':depthList2,
        'holedepths':holeDepthList,
        'chartWidth':chartWidth,
        'chartHeight':chartHeight,
        'datas_1':{'data':phLists[0], 'phase_section':actHoles[0]},
        'datas_2':{'data':phLists[1], 'phase_section':actHoles[1]},
        'datas_3':{'data':phLists[2], 'phase_section':actHoles[2]},
        'datas_4':{'data':phLists[3], 'phase_section':actHoles[3]},
        'datas_5':{'data':phLists[4], 'phase_section':actHoles[4]},
        'datas_6':{'data':phLists[5], 'phase_section':actHoles[5]},
        'datas_7':{'data':phLists[6], 'phase_section':actHoles[6]},
        'timebreakdownlist':timeBreakdownList,
        'timebreakdownsublist1':timeBreakdownSubList1,
        'timebreakdownsublist2':timeBreakdownSubList2
    }

    if memberID == 2: return  render(request, 'rigwell/rig_zzz_eow.html', context)

    return  render(request, 'rigwell/rig_user_eow.html', context)





############################################################################################################## AllWeedJobs ############################################################################################
############################################################################################################## AllWeedJobs ############################################################################################
############################################################################################################## AllWeedJobs ############################################################################################

def isloggedin(request) :
    try:
        if request.session['isloggedin'] == 'yes':
            return 'yes'
        else: return 'no'
    except KeyError:
        return 'no'
    return 'yes'


def mea_jobboard(request):
    if isloggedin(request) == 'yes':
        return render(request, 'meajobs/home.html', {'isloggedin':isloggedin(request)})
    else:
        return render(request, 'meajobs/index.html', {'isloggedin':isloggedin(request)})

def browsejobs(request):
    if isloggedin(request) == 'yes':
        return render(request, 'meajobs/home.html', {'isloggedin':isloggedin(request)})
    else:
        return render(request, 'meajobs/browsejobs.html', {'isloggedin':isloggedin(request)})

def mea_candidates(request):
    return render(request, 'meajobs/candidates.html', {'isloggedin':isloggedin(request)})

def postjob(request):
    return render(request, 'meajobs/new-post.html', {'isloggedin':isloggedin(request)})

def wantjob(request):
    return render(request, 'meajobs/job-post.html', {'isloggedin':isloggedin(request)})

def toblog(request):
    return render(request, 'meajobs/blog.html', {'isloggedin':isloggedin(request)})

def mea_contact(request):
    return render(request, 'meajobs/contact.html', {'isloggedin':isloggedin(request)})

def blogdetail(request):
    return render(request, 'meajobs/blog-single.html', {'isloggedin':isloggedin(request)})

def forgotpassword(request):
    return render(request, 'meajobs/forgot-password.html', {'isloggedin':isloggedin(request)})

def loginpage(request):
    return render(request, 'meajobs/login.html', {'isloggedin':isloggedin(request)})

def signuppage(request):
    return render(request, 'meajobs/signup.html', {'isloggedin':isloggedin(request)})

def jobdetail(request):
    if isloggedin(request) == 'yes':
        return render(request, 'meajobs/job-single.html', {'isloggedin':isloggedin(request)})
    else:
        return render(request, 'meajobs/login.html', {'isloggedin':isloggedin(request)})

def categoryjobs(request):
    category = request.GET['category']
    return render(request, 'meajobs/category-jobs.html', {'category':category, 'isloggedin':isloggedin(request)})

def jobcompanies(request):
    return render(request, 'meajobs/job-companies.html', {'isloggedin':isloggedin(request)})

def companyprofile(request):
    return render(request, 'meajobs/company-profile.html', {'isloggedin':isloggedin(request)})

def companyreviews(request):
    return render(request, 'meajobs/company-reviews.html', {'isloggedin':isloggedin(request)})

def rating(request):
    return render(request, 'meajobs/rating.html', {'isloggedin':isloggedin(request)})

def companysalaries(request):
    return render(request, 'meajobs/company-salaries.html', {'isloggedin':isloggedin(request)})

def companyjobs(request):
    return render(request, 'meajobs/company-jobs.html', {'isloggedin':isloggedin(request)})

def awj_login(request):
    request.session['isloggedin'] = 'yes'
    return render(request, 'meajobs/home.html')

def awj_signup(request):
    request.session['isloggedin'] = 'yes'
    return render(request, 'meajobs/home.html')

def awj_logout(request):
    request.session['isloggedin'] = ''
    return render(request, 'meajobs/index.html', {'isloggedin':isloggedin(request)})


def awj_home(request):
    return render(request, 'meajobs/home.html')



































